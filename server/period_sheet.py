"""Individual Returns by Period sheet creation."""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from constants import BENCHMARK_TICKERS, BENCHMARK_ORDER, CASH_TICKER, FX_TICKER
from market_data import get_price_on_date


def create_period_sheet(wb, df, periods, benchmark_returns, dates, cache):
    """Create the Individual Returns by Period sheet."""
    ws = wb.create_sheet("Period Contributions")
    
    # Set column A width to 10px (empty border column)
    ws.column_dimensions['A'].width = 10 / 7  # Convert pixels to Excel units
    
    # Set column B width for tickers
    ws.column_dimensions['B'].width = 15  # Reasonable width for ticker names
    
    # Set row 1 height to 10px (empty border row)
    ws.row_dimensions[1].height = 10  # Height in points (1 point ≈ 1.33 pixels)
    
    # Freeze columns A and B (Ticker column), but not row 1
    ws.freeze_panes = 'D1'
    
    # Define font
    aptos_font = Font(name='Aptos', size=11)
    aptos_font_bold = Font(name='Aptos', size=11, bold=True)
    
    # Define border styles
    no_border = Border()
    thick_border = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )
    
    # Group periods by month for spacing
    def group_periods_by_month(periods):
        """Group periods by month to determine spacing."""
        month_groups = {}
        for idx, period in enumerate(periods):
            month_key = (period[0].year, period[0].month)
            if month_key not in month_groups:
                month_groups[month_key] = []
            month_groups[month_key].append((idx, period))
        return month_groups
    
    month_groups = group_periods_by_month(periods)
    
    current_row = 2  # Start at row 2, leaving row 1 empty
    
    col_offset = 4  # Periods start at column D (column A is empty, column B is tickers, column C is spacer)
    current_col = col_offset
    
    # Create period headers with month grouping
    for period_idx, period in enumerate(periods):
        # Check if this is the first period in a new month (need more spacing)
        is_new_month = False
        if period_idx == 0:
            is_new_month = True
        else:
            prev_period = periods[period_idx - 1]
            if prev_period[1].month != period[0].month or prev_period[1].year != period[0].year:
                is_new_month = True
        
        # If new month and not first period, add extra spacing
        if is_new_month and period_idx > 0:
            current_col += 1  # Add extra spacer for month separation
        
        start_col = current_col
        end_col = start_col + 2
        start_letter = get_column_letter(start_col)
        end_letter = get_column_letter(end_col)
        
        ws.merge_cells(f'{start_letter}{current_row}:{end_letter}{current_row}')
        period_cell = ws[f'{start_letter}{current_row}']
        start_date_str = period[0].strftime("%d/%m/%Y")
        end_date_str = period[1].strftime("%d/%m/%Y")
        period_cell.value = f"{start_date_str} @ {end_date_str}"
        period_cell.alignment = Alignment(horizontal='center', vertical='center')
        period_cell.font = Font(name='Aptos', size=11, bold=True, color='FFFFFF')
        period_cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        
        # Apply thick border to all cells in merged range for proper border coverage
        for col in range(start_col, end_col + 1):
            cell = ws[f'{get_column_letter(col)}{current_row}']
            cell.border = Border(
                left=Side(style='medium', color='000000') if col == start_col else None,
                right=Side(style='medium', color='000000') if col == end_col else None,
                top=Side(style='medium', color='000000'),
                bottom=Side(style='medium', color='000000')
            )
        
        # Sub-headers - black background, white font
        for sub_idx, sub_header in enumerate(["Weight", "Return", "Contrib."]):
            sub_cell = ws[f'{get_column_letter(start_col + sub_idx)}{current_row + 1}']
            sub_cell.value = sub_header
            sub_cell.font = Font(name='Aptos', size=11, bold=True, color='FFFFFF')
            sub_cell.alignment = Alignment(horizontal='center', vertical='center')
            sub_cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            sub_cell.border = no_border
        
        current_col += 4  # Move to next period (Weight/Return/Contrib/Spacer)
    
    # YTD columns after all periods and spacers
    ytd_col = current_col + 1
    ytd_start_col = ytd_col
    ytd_end_col = ytd_col + 1
    ytd_header_cell = ws[f'{get_column_letter(ytd_col)}{current_row}']
    ws.merge_cells(f'{get_column_letter(ytd_col)}{current_row}:{get_column_letter(ytd_col + 1)}{current_row}')
    ytd_header_cell.value = "Cumulative"
    ytd_header_cell.alignment = Alignment(horizontal='center', vertical='center')
    ytd_header_cell.font = Font(name='Aptos', size=11, bold=True, color='FFFFFF')
    ytd_header_cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    
    # Apply thick border to all cells in merged range for proper border coverage
    for col in range(ytd_start_col, ytd_end_col + 1):
        cell = ws[f'{get_column_letter(col)}{current_row}']
        cell.border = Border(
            left=Side(style='medium', color='000000') if col == ytd_start_col else None,
            right=Side(style='medium', color='000000') if col == ytd_end_col else None,
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
    
    ytd_return_header = ws[f'{get_column_letter(ytd_col)}{current_row + 1}']
    ytd_return_header.value = "YTD Return"
    ytd_return_header.font = Font(name='Aptos', size=11, bold=True, color='FFFFFF')
    ytd_return_header.alignment = Alignment(horizontal='center', vertical='center')
    ytd_return_header.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    ytd_return_header.border = no_border
    
    ytd_contrib_header = ws[f'{get_column_letter(ytd_col + 1)}{current_row + 1}']
    ytd_contrib_header.value = "YTD Contrib."
    ytd_contrib_header.font = Font(name='Aptos', size=11, bold=True, color='FFFFFF')
    ytd_contrib_header.alignment = Alignment(horizontal='center', vertical='center')
    ytd_contrib_header.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    ytd_contrib_header.border = no_border
    
    current_row += 1
    
    # Ticker header on same row as Weight/Return/Contrib headers (column B, not A)
    ticker_header = ws['B' + str(current_row)]
    ticker_header.value = "Ticker"
    ticker_header.font = Font(name='Aptos', size=11, bold=True, color='FFFFFF')
    ticker_header.alignment = Alignment(horizontal='center', vertical='center')
    ticker_header.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    ticker_header.border = thick_border
    
    current_row += 1
    
    prev_weights = {}
    period_col_mapping = {}  # Map period_idx to actual column position
    current_mapped_col = col_offset
    
    # Build column mapping for periods
    for period_idx, period in enumerate(periods):
        is_new_month = False
        if period_idx == 0:
            is_new_month = True
        else:
            prev_period = periods[period_idx - 1]
            if prev_period[1].month != period[0].month or prev_period[1].year != period[0].year:
                is_new_month = True
        
        if is_new_month and period_idx > 0:
            current_mapped_col += 1
        
        period_col_mapping[period_idx] = current_mapped_col
        current_mapped_col += 4
    
    # Data rows start immediately after the ticker header row (no extra row)
    data_start_row = current_row
    holdings_end_row = data_start_row + len(df) - 1
    
    for row_idx, (_, row_data) in enumerate(df.iterrows(), start=data_start_row):
        ticker = row_data['Ticker']
        ticker_cell = ws[f'B{row_idx}']
        ticker_cell.value = ticker
        ticker_cell.border = no_border
        ticker_cell.font = aptos_font
        
        # Alternate row background (very light grey) - skip spacing columns
        # Start with white for first data row (index 0)
        # Use row_idx - data_start_row to get 0-based index for alternating pattern
        data_row_index = row_idx - data_start_row
        if data_row_index % 2 == 1:  # Odd indices (1st, 3rd, 5th...) get grey
            row_fill = PatternFill(start_color='d9d9d9', end_color='d9d9d9', fill_type='solid')
        else:
            row_fill = None
        
        if ticker != CASH_TICKER and ticker not in BENCHMARK_TICKERS.values():
            ticker_cell.font = Font(name='Aptos', size=11, bold=True, italic=True)
        else:
            ticker_cell.font = aptos_font_bold
        
        if row_fill:
            ticker_cell.fill = row_fill
        
        for period_idx, period in enumerate(periods):
            col_idx = period_col_mapping[period_idx]
            weight = row_data.get(f'Weight_{period_idx}', 0.0)
            return_val = row_data.get(f'Return_{period_idx}', 0.0)
            contrib = row_data.get(f'Contrib_{period_idx}', 0.0)
            
            weight_cell = ws[f'{get_column_letter(col_idx)}{row_idx}']
            weight_cell.value = weight
            weight_cell.number_format = '0.00%;(0.00%)'
            weight_cell.alignment = Alignment(horizontal='right', vertical='center')
            weight_cell.border = no_border
            weight_cell.font = aptos_font
            
            # Conditional formatting for weights based on preceding weight
            # Get previous period's weight for this ticker
            prev_weight = None
            if period_idx > 0:
                prev_period = periods[period_idx - 1]
                prev_weight = prev_weights.get(ticker, {}).get(prev_period[0])
            
            # Apply conditional fill (overrides row_fill if weight changed)
            if prev_weight is not None:
                if weight > prev_weight:
                    # Weight increased - green background
                    weight_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    weight_cell.font = Font(name='Aptos', size=11, color='006100')
                elif weight < prev_weight:
                    # Weight decreased - red background
                    weight_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    weight_cell.font = Font(name='Aptos', size=11, color='C00000')
                else:
                    # Weight unchanged - use row fill
                    if row_fill:
                        weight_cell.fill = row_fill
            else:
                # First period or no previous weight - use row fill
                if row_fill:
                    weight_cell.fill = row_fill
            
            if ticker not in prev_weights:
                prev_weights[ticker] = {}
            prev_weights[ticker][period[0]] = weight
            
            return_cell = ws[f'{get_column_letter(col_idx + 1)}{row_idx}']
            return_cell.value = return_val
            return_cell.number_format = '0.00%;(0.00%)'
            return_cell.alignment = Alignment(horizontal='right', vertical='center')
            return_cell.border = no_border
            return_cell.font = aptos_font
            if row_fill:
                return_cell.fill = row_fill
            if return_val >= 0:
                return_cell.font = Font(name='Aptos', size=11, color='006100')
            else:
                return_cell.font = Font(name='Aptos', size=11, color='C00000')
            
            contrib_cell = ws[f'{get_column_letter(col_idx + 2)}{row_idx}']
            contrib_cell.value = contrib
            contrib_cell.number_format = '0.00%;(0.00%)'
            contrib_cell.alignment = Alignment(horizontal='right', vertical='center')
            contrib_cell.border = no_border
            contrib_cell.font = aptos_font
            if row_fill:
                contrib_cell.fill = row_fill
            
            if contrib >= 0:
                contrib_cell.font = Font(name='Aptos', size=11, bold=True, color='006100')
            else:
                contrib_cell.font = Font(name='Aptos', size=11, bold=True, color='C00000')
        
        ytd_return = row_data.get('YTD_Return', 0.0)
        ytd_contrib = row_data.get('YTD_Contrib', 0.0)
        
        ytd_return_cell = ws[f'{get_column_letter(ytd_col)}{row_idx}']
        ytd_return_cell.value = ytd_return
        ytd_return_cell.number_format = '0.00%;(0.00%)'
        ytd_return_cell.alignment = Alignment(horizontal='right', vertical='center')
        ytd_return_cell.border = no_border
        ytd_return_cell.font = aptos_font
        if row_fill:
            ytd_return_cell.fill = row_fill
        if ytd_return >= 0:
            ytd_return_cell.font = Font(name='Aptos', size=11, color='006100')
        else:
            ytd_return_cell.font = Font(name='Aptos', size=11, color='C00000')
        
        ytd_contrib_cell = ws[f'{get_column_letter(ytd_col + 1)}{row_idx}']
        ytd_contrib_cell.value = ytd_contrib
        ytd_contrib_cell.number_format = '0.00%;(0.00%)'
        ytd_contrib_cell.alignment = Alignment(horizontal='right', vertical='center')
        ytd_contrib_cell.border = no_border
        ytd_contrib_cell.font = aptos_font
        if row_fill:
            ytd_contrib_cell.fill = row_fill
        
        if ytd_contrib >= 0:
            ytd_contrib_cell.font = Font(name='Aptos', size=11, bold=True, color='006100')
        else:
            ytd_contrib_cell.font = Font(name='Aptos', size=11, bold=True, color='C00000')
    
    # TOTAL row immediately after last ticker
    total_row = holdings_end_row + 1
    total_ticker_cell = ws[f'B{total_row}']
    total_ticker_cell.value = ""  # Remove "TOTAL" text
    total_ticker_cell.border = no_border
    total_ticker_cell.font = aptos_font_bold
    total_ticker_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # No color fill for TOTAL row
    for period_idx, period in enumerate(periods):
        col_idx = period_col_mapping[period_idx]
        total_weight = df[f'Weight_{period_idx}'].sum()
        total_contrib = df[f'Contrib_{period_idx}'].sum()
        
        weight_cell = ws[f'{get_column_letter(col_idx)}{total_row}']
        weight_cell.value = total_weight
        weight_cell.number_format = '0.00%;(0.00%)'
        weight_cell.alignment = Alignment(horizontal='right', vertical='center')
        weight_cell.font = Font(name='Aptos', size=11, bold=True, color='808080')  # Grey font
        weight_cell.border = no_border
        
        return_cell = ws[f'{get_column_letter(col_idx + 1)}{total_row}']
        return_cell.value = ""
        return_cell.border = no_border
        
        contrib_cell = ws[f'{get_column_letter(col_idx + 2)}{total_row}']
        contrib_cell.value = total_contrib
        contrib_cell.number_format = '0.00%;(0.00%)'
        contrib_cell.alignment = Alignment(horizontal='right', vertical='center')
        contrib_cell.font = aptos_font_bold
        contrib_cell.border = no_border
        
        if total_contrib >= 0:
            contrib_cell.font = Font(name='Aptos', size=11, bold=True, color='006100')
        else:
            contrib_cell.font = Font(name='Aptos', size=11, bold=True, color='C00000')
    
    total_ytd_contrib = df['YTD_Contrib'].sum()
    ytd_return_cell = ws[f'{get_column_letter(ytd_col)}{total_row}']
    ytd_return_cell.value = ""
    ytd_return_cell.border = no_border
    
    ytd_contrib_cell = ws[f'{get_column_letter(ytd_col + 1)}{total_row}']
    ytd_contrib_cell.value = total_ytd_contrib
    ytd_contrib_cell.number_format = '0.00%;(0.00%)'
    ytd_contrib_cell.alignment = Alignment(horizontal='right', vertical='center')
    ytd_contrib_cell.font = aptos_font_bold
    ytd_contrib_cell.border = no_border
    
    if total_ytd_contrib >= 0:
        ytd_contrib_cell.font = Font(name='Aptos', size=11, bold=True, color='006100')
    else:
        ytd_contrib_cell.font = Font(name='Aptos', size=11, bold=True, color='C00000')
    
    # Add spacing row after TOTAL and before benchmarks
    spacing_row = total_row + 1
    ws.row_dimensions[spacing_row].height = 20
    
    # Benchmarks Return header
    benchmark_header_row = spacing_row + 1
    benchmark_header_cell = ws[f'B{benchmark_header_row}']
    benchmark_header_cell.value = "Benchmarks Return are all in Domestic Currencies"
    benchmark_header_cell.font = Font(name='Aptos', size=11, bold=True, color='FF0000')
    benchmark_header_cell.alignment = Alignment(horizontal='left', vertical='center')
    benchmark_header_cell.border = no_border
    
    benchmark_start_row = benchmark_header_row + 1
    
    # Define benchmark colors
    benchmark_colors = {
        "USD/CAD": "fcd5b4",  # Light orange
        "S&P 500": "b8cce4",  # Light blue
        "Dow Jones": "b8cce4",  # Light blue
        "Nasdaq": "b8cce4",  # Light blue
        "ACWI": "b8cce4",  # Light blue
        "TSX60": "d8e4bc"  # Light green
    }
    
    for bench_name in BENCHMARK_ORDER:
        bench_row = benchmark_start_row
        bench_ticker_cell = ws[f'B{bench_row}']
        bench_ticker_cell.value = bench_name
        bench_ticker_cell.font = aptos_font_bold
        bench_ticker_cell.border = no_border
        bench_ticker_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Fill benchmark row with color (skip spacing columns)
        bench_color = benchmark_colors.get(bench_name, "FFFFFF")
        bench_fill = PatternFill(start_color=bench_color, end_color=bench_color, fill_type='solid')
        bench_ticker_cell.fill = bench_fill
        
        for period_idx, period in enumerate(periods):
            col_idx = period_col_mapping[period_idx]
            bench_return = benchmark_returns[bench_name].get(period, 0.0)
            
            # Weight column (empty)
            ws[f'{get_column_letter(col_idx)}{bench_row}'].value = ""
            ws[f'{get_column_letter(col_idx)}{bench_row}'].border = no_border
            ws[f'{get_column_letter(col_idx)}{bench_row}'].fill = bench_fill
            
            # Return column
            return_cell = ws[f'{get_column_letter(col_idx + 1)}{bench_row}']
            return_cell.value = bench_return
            return_cell.number_format = '0.00%;(0.00%)'
            return_cell.alignment = Alignment(horizontal='right', vertical='center')
            return_cell.border = no_border
            return_cell.font = aptos_font
            return_cell.fill = bench_fill
            if bench_return >= 0:
                return_cell.font = Font(name='Aptos', size=11, color='006100')
            else:
                return_cell.font = Font(name='Aptos', size=11, color='C00000')
            
            # Contrib column (empty)
            ws[f'{get_column_letter(col_idx + 2)}{bench_row}'].value = ""
            ws[f'{get_column_letter(col_idx + 2)}{bench_row}'].border = no_border
            ws[f'{get_column_letter(col_idx + 2)}{bench_row}'].fill = bench_fill
        
        first_date = pd.to_datetime(dates[0], format="%d/%m/%Y")
        last_date = pd.to_datetime(dates[-1], format="%d/%m/%Y")
        
        ticker = BENCHMARK_TICKERS[bench_name]
        if ticker == FX_TICKER:
            fx_start = get_price_on_date(FX_TICKER, first_date, cache)
            fx_end = get_price_on_date(FX_TICKER, last_date, cache)
            ytd_bench_return = (fx_end / fx_start) - 1
        else:
            price_start = get_price_on_date(ticker, first_date, cache)
            price_end = get_price_on_date(ticker, last_date, cache)
            ytd_bench_return = (price_end / price_start) - 1
        
        # YTD Return
        ytd_return_cell = ws[f'{get_column_letter(ytd_col)}{bench_row}']
        ytd_return_cell.value = ytd_bench_return
        ytd_return_cell.number_format = '0.00%;(0.00%)'
        ytd_return_cell.alignment = Alignment(horizontal='right', vertical='center')
        ytd_return_cell.border = no_border
        ytd_return_cell.font = aptos_font
        ytd_return_cell.fill = bench_fill
        if ytd_bench_return >= 0:
            ytd_return_cell.font = Font(name='Aptos', size=11, color='006100')
        else:
            ytd_return_cell.font = Font(name='Aptos', size=11, color='C00000')
        
        # YTD Contrib (empty)
        ws[f'{get_column_letter(ytd_col + 1)}{bench_row}'].value = ""
        ws[f'{get_column_letter(ytd_col + 1)}{bench_row}'].border = no_border
        ws[f'{get_column_letter(ytd_col + 1)}{bench_row}'].fill = bench_fill
        
        benchmark_start_row += 1
    
    # Set column widths - including spacers
    # Calculate max column needed
    max_col = max([ytd_col + 1] + [period_col_mapping[idx] + 3 for idx in period_col_mapping])
    
    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        
        if col == 1:  # Column A - empty border column (10px)
            ws.column_dimensions[col_letter].width = 10 / 7  # 10px border column
        elif col == 2:  # Ticker column (Column B)
            ws.column_dimensions[col_letter].width = 15  # Reasonable width for ticker names
        elif col == 3:  # Spacer column (Column C)
            ws.column_dimensions[col_letter].width = 10 / 7  # 10px spacer column
        elif col >= col_offset and col < ytd_col:
            # Check if this is a spacer column
            is_spacer = False
            for period_idx, mapped_col in period_col_mapping.items():
                # Spacer is at mapped_col + 3
                if col == mapped_col + 3:
                    is_spacer = True
                    break
                # Also check for month separators (extra spacers)
                if period_idx + 1 < len(periods):
                    next_period_start = period_col_mapping[period_idx + 1]
                    if col == next_period_start - 1:
                        is_spacer = True
                        break
            
            if is_spacer:
                ws.column_dimensions[col_letter].width = 20 / 7  # ~20 pixels
            else:
                # Check if it's within a period group (Weight/Return/Contrib)
                for period_idx, mapped_col in period_col_mapping.items():
                    if mapped_col <= col < mapped_col + 3:
                        ws.column_dimensions[col_letter].width = 65 / 7  # ~65 pixels
                        break
        elif col >= ytd_col:
            ws.column_dimensions[col_letter].width = 100 / 7  # YTD columns 85px each
        else:
            ws.column_dimensions[col_letter].width = 12  # Default for any other columns
    
    # Turn off grid lines
    ws.sheet_view.showGridLines = False
    
    # Apply thick borders around ticker column (Column B)
    # Only around ticker data cells, NOT the header or TOTAL row
    ticker_data_start_row = data_start_row
    ticker_data_end_row = holdings_end_row
    
    # Apply borders only to ticker data cells (not header or TOTAL)
    for row in range(ticker_data_start_row, ticker_data_end_row + 1):
        ticker_cell = ws[f'B{row}']
        # Top border on first data row, bottom border on last data row, left/right on all data rows
        top_border = Side(style='medium', color='000000') if row == ticker_data_start_row else None
        bottom_border = Side(style='medium', color='000000') if row == ticker_data_end_row else None
        ticker_cell.border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=top_border,
            bottom=bottom_border
        )
    
    # Apply thick borders around each period table
    header_row = data_start_row - 2  # Date span row
    subheader_row = data_start_row - 1  # Weight/Return/Contrib row (and Ticker header)
    # data_start_row already defined above
    data_end_row = holdings_end_row  # Last ticker row
    
    for period_idx, period in enumerate(periods):
        col_idx = period_col_mapping[period_idx]
        
        # Thick border on date span merged cell - apply to all cells in merged range
        end_col = col_idx + 2
        for col in range(col_idx, end_col + 1):
            cell = ws[f'{get_column_letter(col)}{header_row}']
            cell.border = Border(
                left=Side(style='medium', color='000000') if col == col_idx else None,
                right=Side(style='medium', color='000000') if col == end_col else None,
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
        
        # Top border (subheader row)
        for col in range(col_idx, col_idx + 3):
            cell = ws[f'{get_column_letter(col)}{subheader_row}']
            current_border = cell.border
            cell.border = Border(
                left=Side(style='medium', color='000000') if col == col_idx else None,
                right=Side(style='medium', color='000000') if col == col_idx + 2 else None,
                top=Side(style='medium', color='000000'),
                bottom=None
            )
        
        # Left and right borders for data rows only (NOT TOTAL row)
        for row in range(data_start_row, holdings_end_row + 1):
            # Left border (first column of period)
            left_cell = ws[f'{get_column_letter(col_idx)}{row}']
            current_border = left_cell.border
            new_left_border = Border(
                left=Side(style='medium', color='000000'),
                right=current_border.right if current_border and current_border.right else None,
                top=Side(style='medium', color='000000') if row == data_start_row else None,
                bottom=Side(style='medium', color='000000') if row == holdings_end_row else None
            )
            left_cell.border = new_left_border
            
            # Right border (last column of period)
            right_cell = ws[f'{get_column_letter(col_idx + 2)}{row}']
            current_border = right_cell.border
            new_right_border = Border(
                left=current_border.left if current_border and current_border.left else None,
                right=Side(style='medium', color='000000'),
                top=Side(style='medium', color='000000') if row == data_start_row else None,
                bottom=Side(style='medium', color='000000') if row == holdings_end_row else None
            )
            right_cell.border = new_right_border
        
        # Bottom border (last data row, not TOTAL row)
        for col in range(col_idx, col_idx + 3):
            cell = ws[f'{get_column_letter(col)}{holdings_end_row}']
            current_border = cell.border
            cell.border = Border(
                left=Side(style='medium', color='000000') if col == col_idx else (current_border.left if current_border and current_border.left else None),
                right=Side(style='medium', color='000000') if col == col_idx + 2 else (current_border.right if current_border and current_border.right else None),
                top=current_border.top if current_border and current_border.top else None,
                bottom=Side(style='medium', color='000000')
            )
    
    # Apply thick borders around YTD/Cumulative columns (same format as period tables)
    # Top border (Cumulative merged cell) - apply to all cells in merged range
    ytd_end_col = ytd_col + 1
    for col in range(ytd_col, ytd_end_col + 1):
        cell = ws[f'{get_column_letter(col)}{header_row}']
        cell.border = Border(
            left=Side(style='medium', color='000000') if col == ytd_col else None,
            right=Side(style='medium', color='000000') if col == ytd_end_col else None,
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )
    
    # Top border (subheader row - YTD Return and YTD Contrib.)
    for col in range(ytd_col, ytd_col + 2):
        cell = ws[f'{get_column_letter(col)}{subheader_row}']
        current_border = cell.border
        cell.border = Border(
            left=Side(style='medium', color='000000') if col == ytd_col else None,
            right=Side(style='medium', color='000000') if col == ytd_col + 1 else None,
            top=Side(style='medium', color='000000'),
            bottom=None
        )
    
    # Left and right borders for data rows only (NOT TOTAL row)
    for row in range(data_start_row, holdings_end_row + 1):
        # Left border (YTD Return column)
        left_cell = ws[f'{get_column_letter(ytd_col)}{row}']
        current_border = left_cell.border
        new_left_border = Border(
            left=Side(style='medium', color='000000'),
            right=current_border.right if current_border and current_border.right else None,
            top=Side(style='medium', color='000000') if row == data_start_row else None,
            bottom=Side(style='medium', color='000000') if row == holdings_end_row else None
        )
        left_cell.border = new_left_border
        
        # Right border (YTD Contrib column)
        right_cell = ws[f'{get_column_letter(ytd_col + 1)}{row}']
        current_border = right_cell.border
        new_right_border = Border(
            left=current_border.left if current_border and current_border.left else None,
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000') if row == data_start_row else None,
            bottom=Side(style='medium', color='000000') if row == holdings_end_row else None
        )
        right_cell.border = new_right_border
    
    # Bottom border (last data row, not TOTAL row)
    for col in range(ytd_col, ytd_col + 2):
        cell = ws[f'{get_column_letter(col)}{holdings_end_row}']
        current_border = cell.border
        cell.border = Border(
            left=Side(style='medium', color='000000') if col == ytd_col else (current_border.left if current_border and current_border.left else None),
            right=Side(style='medium', color='000000') if col == ytd_col + 1 else (current_border.right if current_border and current_border.right else None),
            top=current_border.top if current_border and current_border.top else None,
            bottom=Side(style='medium', color='000000')
        )
    
    # Set print settings - center vertically and horizontally
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    
    # Set custom header and footer
    ws.oddHeader.center.text = '&"Aptos,Bold"&26&[Tab]'
    ws.oddFooter.center.text = '&"Aptos,Bold"&12Page &[Page] of &[Pages]'

