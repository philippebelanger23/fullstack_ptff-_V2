"""Monthly contributions sheet creation and related functions."""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
import pandas as pd
import datetime
from constants import BENCHMARK_TICKERS, BENCHMARK_ORDER, CASH_TICKER, FX_TICKER
from market_data import get_price_on_date


def create_monthly_periods(dates, periods=None):
    """Create monthly periods from dates - use actual periods to determine month boundaries."""
    monthly_periods = []
    
    if periods is None or not periods:
        # Fallback to dates-based logic if periods not provided
        parsed_dates = [pd.to_datetime(d, format="%d/%m/%Y") for d in dates]
        
        if not parsed_dates:
            return monthly_periods
        
        # Group dates by calendar month
        month_groups = {}
        for date in parsed_dates:
            month_key = (date.year, date.month)
            if month_key not in month_groups:
                month_groups[month_key] = []
            month_groups[month_key].append(date)
        
        # Find the first January month (skip December of previous year)
        sorted_months = sorted(month_groups.keys())
        first_january_idx = None
        for i, (year, month) in enumerate(sorted_months):
            if month == 1:  # January
                first_january_idx = i
                break
        
        # If no January found, start from first month
        if first_january_idx is None:
            first_january_idx = 0
        
        # Process months starting from January
        sorted_months = sorted_months[first_january_idx:]
        
        # For each month, find the actual start and end dates
        for i, month_key in enumerate(sorted_months):
            year, month = month_key
            month_dates = sorted(month_groups[month_key])
            
            if len(month_dates) < 1:
                continue
            
            # Find start date - use previous month's last date if available, otherwise first date of this month
            if i > 0:
                prev_month_key = sorted_months[i - 1]
                prev_month_dates = sorted(month_groups[prev_month_key])
                if prev_month_dates:
                    month_start = prev_month_dates[-1]
                else:
                    month_start = month_dates[0]
            else:
                month_start = month_dates[0]
            
            month_end = month_dates[-1]
            monthly_periods.append((month_start, month_end))
    else:
        # Use periods to determine month boundaries (more accurate)
        # Group periods by calendar month
        month_periods = {}
        for period in periods:
            period_start, period_end = period
            # Use period_end to determine which month it belongs to
            month_key = (period_end.year, period_end.month)
            if month_key not in month_periods:
                month_periods[month_key] = []
            month_periods[month_key].append(period)
        
        # Find the first January month
        sorted_months = sorted(month_periods.keys())
        first_january_idx = None
        for i, (year, month) in enumerate(sorted_months):
            if month == 1:  # January
                first_january_idx = i
                break
        
        if first_january_idx is None:
            first_january_idx = 0
        
        sorted_months = sorted_months[first_january_idx:]
        
        # For each month, find the actual start and end dates from periods
        for i, month_key in enumerate(sorted_months):
            year, month = month_key
            month_period_list = month_periods[month_key]
            
            if len(month_period_list) < 1:
                continue
            
            # Find the earliest start and latest end in this month's periods
            month_starts = [p[0] for p in month_period_list]
            month_ends = [p[1] for p in month_period_list]
            
            # Start date: use previous month's last period end if available, otherwise first period start of this month
            if i > 0:
                prev_month_key = sorted_months[i - 1]
                prev_month_periods = month_periods[prev_month_key]
                if prev_month_periods:
                    prev_month_end = max(p[1] for p in prev_month_periods)
                    month_start = prev_month_end
                else:
                    month_start = min(month_starts)
            else:
                # First month (January) - use first period start
                month_start = min(month_starts)
            
            # End date is the last period end in this month
            month_end = max(month_ends)
            
            monthly_periods.append((month_start, month_end))
    
    return monthly_periods


def calculate_monthly_returns(weights_dict, nav_dict, monthly_periods, prices, cache):
    """Calculate returns for monthly periods."""
    from constants import CASH_TICKER, FX_TICKER
    from market_data import get_fx_return, get_price_on_date
    
    all_tickers = sorted(weights_dict.keys())
    monthly_returns = {}
    
    for ticker in all_tickers:
        if ticker == CASH_TICKER:
            monthly_returns[ticker] = {}
            for period in monthly_periods:
                monthly_returns[ticker][period] = 0.0
            continue
        
        monthly_returns[ticker] = {}
        for period in monthly_periods:
            start_date, end_date = period
            
            # Get prices for start and end dates
            if ticker in nav_dict:
                if start_date in nav_dict[ticker] and end_date in nav_dict[ticker]:
                    price_start = nav_dict[ticker][start_date]
                    price_end = nav_dict[ticker][end_date]
                elif start_date in prices.get(ticker, {}) and end_date in prices.get(ticker, {}):
                    price_start = prices[ticker][start_date]
                    price_end = prices[ticker][end_date]
                else:
                    price_start = get_price_on_date(ticker, start_date, cache)
                    price_end = get_price_on_date(ticker, end_date, cache)
            else:
                if start_date in prices.get(ticker, {}) and end_date in prices.get(ticker, {}):
                    price_start = prices[ticker][start_date]
                    price_end = prices[ticker][end_date]
                else:
                    price_start = get_price_on_date(ticker, start_date, cache)
                    price_end = get_price_on_date(ticker, end_date, cache)
            
            period_return = (price_end / price_start) - 1
            
            # Mutual funds (in nav_dict) use NAV data directly without FX adjustment
            if ticker in nav_dict:
                monthly_returns[ticker][period] = period_return
            elif ticker.endswith('.TO') or ticker == "^GSPTSE":
                monthly_returns[ticker][period] = period_return
            else:
                fx_return = get_fx_return(start_date, end_date, cache)
                cad_adjusted_return = (1 + period_return) * (1 + fx_return) - 1
                monthly_returns[ticker][period] = cad_adjusted_return
    
    return monthly_returns


def calculate_monthly_benchmark_returns(monthly_periods, cache):
    """Calculate benchmark returns for monthly periods."""
    from constants import BENCHMARK_TICKERS, FX_TICKER
    from market_data import get_fx_return, get_price_on_date
    
    benchmark_returns = {}
    
    for bench_name, ticker in BENCHMARK_TICKERS.items():
        benchmark_returns[bench_name] = {}
        for period in monthly_periods:
            start_date, end_date = period
            
            if ticker == FX_TICKER:
                benchmark_returns[bench_name][period] = get_fx_return(start_date, end_date, cache)
            else:
                price_start = get_price_on_date(ticker, start_date, cache)
                price_end = get_price_on_date(ticker, end_date, cache)
                benchmark_returns[bench_name][period] = (price_end / price_start) - 1
    
    return benchmark_returns


def build_monthly_dataframe(weights_dict, monthly_returns, monthly_periods, dates, cache, nav_dict=None, periods=None, period_df=None):
    """Build monthly results DataFrame - aggregate data from period sheet only."""
    from constants import CASH_TICKER, FX_TICKER
    from market_data import get_price_on_date, get_fx_return
    
    if nav_dict is None:
        nav_dict = {}
    
    if periods is None or period_df is None:
        raise ValueError("periods and period_df must be provided to build monthly dataframe")
    
    all_tickers = sorted(weights_dict.keys())
    
    data = []
    for ticker in all_tickers:
        row = {"Ticker": ticker}
        
        for period_idx, monthly_period in enumerate(monthly_periods):
            monthly_start, monthly_end = monthly_period
            
            # Calculate monthly return from start to end of month (actual return, not weighted)
            monthly_return = 0.0
            if ticker == CASH_TICKER:
                monthly_return = 0.0
            else:
                # Get prices for start and end of month
                if ticker in nav_dict:
                    if monthly_start in nav_dict[ticker] and monthly_end in nav_dict[ticker]:
                        price_start = nav_dict[ticker][monthly_start]
                        price_end = nav_dict[ticker][monthly_end]
                    else:
                        price_start = get_price_on_date(ticker, monthly_start, cache)
                        price_end = get_price_on_date(ticker, monthly_end, cache)
                else:
                    price_start = get_price_on_date(ticker, monthly_start, cache)
                    price_end = get_price_on_date(ticker, monthly_end, cache)
                
                if price_start and price_start > 0:
                    period_return = (price_end / price_start) - 1
                    
                    # Apply FX adjustment if needed
                    if ticker in nav_dict:
                        monthly_return = period_return
                    elif ticker.endswith('.TO') or ticker == "^GSPTSE":
                        monthly_return = period_return
                    else:
                        fx_return = get_fx_return(monthly_start, monthly_end, cache)
                        monthly_return = (1 + period_return) * (1 + fx_return) - 1
                else:
                    monthly_return = 0.0
            
            # Sum contributions from all subperiods within this month from period_sheet
            contribution = 0.0
            ticker_data = period_df[period_df['Ticker'] == ticker]
            
            if not ticker_data.empty:
                # Find all periods that fall within this monthly period
                for period_idx_sub, period in enumerate(periods):
                    period_start, period_end = period
                    # Check if this period falls within the monthly period
                    # Period should be completely within the monthly period
                    if period_start >= monthly_start and period_end <= monthly_end:
                        contrib_col = f'Contrib_{period_idx_sub}'
                        if contrib_col in period_df.columns:
                            contrib_val = ticker_data[contrib_col].iloc[0]
                            if pd.notna(contrib_val):
                                contribution += contrib_val
            
            row[f"Return_{period_idx}"] = monthly_return
            row[f"Contrib_{period_idx}"] = contribution
        
        # Calculate YTD return and contribution
        if ticker == CASH_TICKER:
            ytd_return = 0.0
            ytd_contrib = 0.0
        else:
            # Calculate YTD return from first to last date of all monthly periods
            if monthly_periods:
                first_period = monthly_periods[0]
                last_period = monthly_periods[-1]
                first_date = first_period[0]
                last_date = last_period[1]
                
                # Get prices for first and last dates
                if ticker in nav_dict:
                    if first_date in nav_dict[ticker] and last_date in nav_dict[ticker]:
                        price_start = nav_dict[ticker][first_date]
                        price_end = nav_dict[ticker][last_date]
                    else:
                        price_start = get_price_on_date(ticker, first_date, cache)
                        price_end = get_price_on_date(ticker, last_date, cache)
                else:
                    price_start = get_price_on_date(ticker, first_date, cache)
                    price_end = get_price_on_date(ticker, last_date, cache)
                
                if price_start and price_start > 0:
                    ytd_return = (price_end / price_start) - 1
                    
                    # Apply FX adjustment if needed
                    if not (ticker.endswith('.TO') or ticker == "^GSPTSE" or ticker in nav_dict):
                        fx_return = get_fx_return(first_date, last_date, cache)
                        ytd_return = (1 + ytd_return) * (1 + fx_return) - 1
                else:
                    ytd_return = 0.0
            else:
                ytd_return = 0.0
            
            # YTD contribution is sum of all monthly contributions
            ytd_contrib = sum(
                row.get(f'Contrib_{period_idx}', 0.0)
                for period_idx in range(len(monthly_periods))
            )
        
        row["YTD_Return"] = ytd_return
        row["YTD_Contrib"] = ytd_contrib
        
        data.append(row)
    
    df = pd.DataFrame(data)
    
    if not df.empty and "YTD_Contrib" in df.columns:
        df = df.sort_values("YTD_Contrib", ascending=False)
    
    return df


def create_monthly_sheet(wb, df, monthly_periods, monthly_benchmark_returns, dates, cache):
    """Create the monthly contributions sheet."""
    ws = wb.create_sheet("Monthly Contributions")
    
    # Set column A width to 10px (empty border column)
    ws.column_dimensions['A'].width = 10 / 7  # Convert pixels to Excel units
    
    # Set column B width for tickers
    ws.column_dimensions['B'].width = 15  # Reasonable width for ticker names
    
    # Set row 1 height to 10px (empty border row)
    ws.row_dimensions[1].height = 10  # Height in points (1 point ≈ 1.33 pixels)
    
    # Freeze columns A and B (Ticker column), but not row 1
    ws.freeze_panes = 'D1'
    
    # Define font
    aptos_font = Font(name='Aptos', size=11)
    aptos_font_bold = Font(name='Aptos', size=11, bold=True)
    
    # Define border styles
    no_border = Border()
    thick_border = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )
    
    current_row = 2  # Start at row 2, leaving row 1 empty
    col_offset = 4  # Periods start at column D (column A is empty, column B is tickers, column C is spacer)
    current_col = col_offset
    
    # Create monthly period headers
    # Ensure we have 12 months (create empty columns for October, November, December if needed)
    months_to_create = 12
    for month_idx in range(months_to_create):
        if month_idx < len(monthly_periods):
            period = monthly_periods[month_idx]
            month_name = period[1].strftime("%B")
        else:
            # Create empty month columns for months without data
            # Calculate month based on first period's year
            if monthly_periods:
                first_year = monthly_periods[0][1].year
                month_date = datetime.date(first_year, month_idx + 1, 1)
                month_name = month_date.strftime("%B")
            else:
                month_name = ["January", "February", "March", "April", "May", "June",
                              "July", "August", "September", "October", "November", "December"][month_idx]
        
        start_col = current_col
        end_col = start_col + 1  # Only Return and Contrib (no Weight)
        start_letter = get_column_letter(start_col)
        end_letter = get_column_letter(end_col)
        
        ws.merge_cells(f'{start_letter}{current_row}:{end_letter}{current_row}')
        period_cell = ws[f'{start_letter}{current_row}']
        period_cell.value = month_name
        period_cell.alignment = Alignment(horizontal='center', vertical='center')
        period_cell.font = Font(name='Aptos', size=11, bold=True, color='FFFFFF')
        period_cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        
        # Apply thick border to all cells in merged range
        for col in range(start_col, end_col + 1):
            cell = ws[f'{get_column_letter(col)}{current_row}']
            cell.border = Border(
                left=Side(style='medium', color='000000') if col == start_col else None,
                right=Side(style='medium', color='000000') if col == end_col else None,
                top=Side(style='medium', color='000000'),
                bottom=Side(style='medium', color='000000')
            )
        
        # Sub-headers (only Return and Contrib, no Weight)
        for sub_idx, sub_header in enumerate(["Return", "Contrib."]):
            sub_cell = ws[f'{get_column_letter(start_col + sub_idx)}{current_row + 1}']
            sub_cell.value = sub_header
            sub_cell.font = Font(name='Aptos', size=11, bold=True, color='FFFFFF')
            sub_cell.alignment = Alignment(horizontal='center', vertical='center')
            sub_cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            sub_cell.border = no_border
        
        current_col += 3  # Move to next period (Return/Contrib/Spacer)
    
    # YTD columns
    ytd_col = current_col + 1
    ytd_start_col = ytd_col
    ytd_end_col = ytd_col + 1
    ytd_header_cell = ws[f'{get_column_letter(ytd_col)}{current_row}']
    ws.merge_cells(f'{get_column_letter(ytd_col)}{current_row}:{get_column_letter(ytd_col + 1)}{current_row}')
    ytd_header_cell.value = "Cumulative"
    ytd_header_cell.alignment = Alignment(horizontal='center', vertical='center')
    ytd_header_cell.font = Font(name='Aptos', size=11, bold=True, color='FFFFFF')
    ytd_header_cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    
    # Apply thick border to all cells in merged range
    for col in range(ytd_start_col, ytd_end_col + 1):
        cell = ws[f'{get_column_letter(col)}{current_row}']
        cell.border = Border(
            left=Side(style='medium', color='000000') if col == ytd_start_col else None,
            right=Side(style='medium', color='000000') if col == ytd_end_col else None,
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
    
    ytd_return_header = ws[f'{get_column_letter(ytd_col)}{current_row + 1}']
    ytd_return_header.value = "YTD Return"
    ytd_return_header.font = Font(name='Aptos', size=11, bold=True, color='FFFFFF')
    ytd_return_header.alignment = Alignment(horizontal='center', vertical='center')
    ytd_return_header.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    ytd_return_header.border = no_border
    
    ytd_contrib_header = ws[f'{get_column_letter(ytd_col + 1)}{current_row + 1}']
    ytd_contrib_header.value = "YTD Contrib."
    ytd_contrib_header.font = Font(name='Aptos', size=11, bold=True, color='FFFFFF')
    ytd_contrib_header.alignment = Alignment(horizontal='center', vertical='center')
    ytd_contrib_header.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    ytd_contrib_header.border = no_border
    
    current_row += 1
    
    # Ticker header on same row as Return/Contrib headers (column B, not A)
    ticker_header = ws['B' + str(current_row)]
    ticker_header.value = "Ticker"
    ticker_header.font = Font(name='Aptos', size=11, bold=True, color='FFFFFF')
    ticker_header.alignment = Alignment(horizontal='center', vertical='center')
    ticker_header.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    ticker_header.border = thick_border
    
    current_row += 1
    
    # Build column mapping for monthly periods (no weight column)
    # Create mapping for all 12 months, even if some don't have data
    period_col_mapping = {}
    current_mapped_col = col_offset
    for month_idx in range(12):
        period_col_mapping[month_idx] = current_mapped_col
        current_mapped_col += 3  # Return/Contrib/Spacer
    
    # Add column page breaks after March, June, September, and December
    # Page breaks: after March (idx 2), June (idx 5), September (idx 8), December (idx 11)
    break_points = [3, 6, 9, 12]  # April, July, October, January (next year) - breaks after previous month
    for break_idx in break_points:
        # Break before the spacer column (at the end of Contrib column, not at spacer)
        prev_month_idx = break_idx - 1
        if prev_month_idx >= 0 and prev_month_idx in period_col_mapping:
            # Break at the end of previous month's Contrib column (Return + Contrib = +1)
            break_col = period_col_mapping[prev_month_idx] + 1
            ws.col_breaks.append(Break(id=break_col))
    
    # Set print titles to repeat ticker column (B) on the left
    ws.print_title_cols = 'B:B'
    
    data_start_row = current_row
    holdings_end_row = data_start_row + len(df) - 1
    
    # Data rows
    for row_idx, (_, row_data) in enumerate(df.iterrows(), start=data_start_row):
        ticker = row_data['Ticker']
        ticker_cell = ws[f'B{row_idx}']
        ticker_cell.value = ticker
        ticker_cell.border = no_border
        ticker_cell.font = aptos_font
        
        # Alternate row background
        data_row_index = row_idx - data_start_row
        if data_row_index % 2 == 1:
            row_fill = PatternFill(start_color='d9d9d9', end_color='d9d9d9', fill_type='solid')
        else:
            row_fill = None
        
        if ticker != CASH_TICKER and ticker not in BENCHMARK_TICKERS.values():
            ticker_cell.font = Font(name='Aptos', size=11, bold=True, italic=True)
        else:
            ticker_cell.font = aptos_font_bold
        
        if row_fill:
            ticker_cell.fill = row_fill
        
        for period_idx in range(12):  # All 12 months
            col_idx = period_col_mapping[period_idx]
            # Get data if period exists, otherwise use 0.0
            has_data = period_idx < len(monthly_periods)
            if has_data:
                return_val = row_data.get(f'Return_{period_idx}', 0.0)
                contrib = row_data.get(f'Contrib_{period_idx}', 0.0)
            else:
                return_val = 0.0
                contrib = 0.0
            
            return_cell = ws[f'{get_column_letter(col_idx)}{row_idx}']
            if has_data:
                return_cell.value = return_val
                return_cell.number_format = '0.00%;(0.00%)'
                return_cell.alignment = Alignment(horizontal='right', vertical='center')
                if return_val >= 0:
                    return_cell.font = Font(name='Aptos', size=11, color='006100')
                else:
                    return_cell.font = Font(name='Aptos', size=11, color='C00000')
            else:
                return_cell.value = "No Data"
                return_cell.number_format = '@'  # Text format
                return_cell.alignment = Alignment(horizontal='center', vertical='center')
                return_cell.font = Font(name='Aptos', size=11, color='808080')  # Grey font
                return_cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # Light grey background
            return_cell.border = no_border
            if row_fill and has_data:
                return_cell.fill = row_fill
            
            contrib_cell = ws[f'{get_column_letter(col_idx + 1)}{row_idx}']
            if has_data:
                contrib_cell.value = contrib
                contrib_cell.number_format = '0.00%;(0.00%)'
                contrib_cell.alignment = Alignment(horizontal='right', vertical='center')
                if contrib >= 0:
                    contrib_cell.font = Font(name='Aptos', size=11, bold=True, color='006100')
                else:
                    contrib_cell.font = Font(name='Aptos', size=11, bold=True, color='C00000')
            else:
                contrib_cell.value = "No Data"
                contrib_cell.number_format = '@'  # Text format
                contrib_cell.alignment = Alignment(horizontal='center', vertical='center')
                contrib_cell.font = Font(name='Aptos', size=11, color='808080')  # Grey font (not bold)
                contrib_cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # Light grey background
            contrib_cell.border = no_border
            if row_fill and has_data:
                contrib_cell.fill = row_fill
        
        ytd_return = row_data.get('YTD_Return', 0.0)
        ytd_contrib = row_data.get('YTD_Contrib', 0.0)
        
        ytd_return_cell = ws[f'{get_column_letter(ytd_col)}{row_idx}']
        ytd_return_cell.value = ytd_return
        ytd_return_cell.number_format = '0.00%;(0.00%)'
        ytd_return_cell.alignment = Alignment(horizontal='right', vertical='center')
        ytd_return_cell.border = no_border
        ytd_return_cell.font = aptos_font
        if row_fill:
            ytd_return_cell.fill = row_fill
        if ytd_return >= 0:
            ytd_return_cell.font = Font(name='Aptos', size=11, color='006100')
        else:
            ytd_return_cell.font = Font(name='Aptos', size=11, color='C00000')
        
        ytd_contrib_cell = ws[f'{get_column_letter(ytd_col + 1)}{row_idx}']
        ytd_contrib_cell.value = ytd_contrib
        ytd_contrib_cell.number_format = '0.00%;(0.00%)'
        ytd_contrib_cell.alignment = Alignment(horizontal='right', vertical='center')
        ytd_contrib_cell.border = no_border
        ytd_contrib_cell.font = aptos_font
        if row_fill:
            ytd_contrib_cell.fill = row_fill
        
        if ytd_contrib >= 0:
            ytd_contrib_cell.font = Font(name='Aptos', size=11, bold=True, color='006100')
        else:
            ytd_contrib_cell.font = Font(name='Aptos', size=11, bold=True, color='C00000')
    
    # TOTAL row
    total_row = holdings_end_row + 1
    total_ticker_cell = ws[f'B{total_row}']
    total_ticker_cell.value = ""
    total_ticker_cell.border = no_border
    total_ticker_cell.font = aptos_font_bold
    total_ticker_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for period_idx in range(12):  # All 12 months
        col_idx = period_col_mapping[period_idx]
        # Get total contrib if period exists, otherwise use 0.0
        has_data = period_idx < len(monthly_periods) and f'Contrib_{period_idx}' in df.columns
        if has_data:
            total_contrib = df[f'Contrib_{period_idx}'].sum()
        else:
            total_contrib = 0.0
        
        return_cell = ws[f'{get_column_letter(col_idx)}{total_row}']
        return_cell.value = ""
        return_cell.border = no_border
        
        contrib_cell = ws[f'{get_column_letter(col_idx + 1)}{total_row}']
        if has_data:
            contrib_cell.value = total_contrib
            contrib_cell.number_format = '0.00%;(0.00%)'
            contrib_cell.alignment = Alignment(horizontal='right', vertical='center')
            if total_contrib >= 0:
                contrib_cell.font = Font(name='Aptos', size=11, bold=True, color='006100')
            else:
                contrib_cell.font = Font(name='Aptos', size=11, bold=True, color='C00000')
        else:
            contrib_cell.value = "No Data"
            contrib_cell.number_format = '@'  # Text format
            contrib_cell.alignment = Alignment(horizontal='center', vertical='center')
            contrib_cell.font = Font(name='Aptos', size=11, color='808080')  # Grey font (not bold)
            contrib_cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # Light grey background
        contrib_cell.border = no_border
    
    total_ytd_contrib = df['YTD_Contrib'].sum()
    ytd_return_cell = ws[f'{get_column_letter(ytd_col)}{total_row}']
    ytd_return_cell.value = ""
    ytd_return_cell.border = no_border
    
    ytd_contrib_cell = ws[f'{get_column_letter(ytd_col + 1)}{total_row}']
    ytd_contrib_cell.value = total_ytd_contrib
    ytd_contrib_cell.number_format = '0.00%;(0.00%)'
    ytd_contrib_cell.alignment = Alignment(horizontal='right', vertical='center')
    ytd_contrib_cell.font = aptos_font_bold
    ytd_contrib_cell.border = no_border
    
    if total_ytd_contrib >= 0:
        ytd_contrib_cell.font = Font(name='Aptos', size=11, bold=True, color='006100')
    else:
        ytd_contrib_cell.font = Font(name='Aptos', size=11, bold=True, color='C00000')
    
    # Spacing row before benchmarks
    spacing_row = total_row + 1
    ws.row_dimensions[spacing_row].height = 20
    
    # Benchmarks Return header
    benchmark_header_row = spacing_row + 1
    benchmark_header_cell = ws[f'B{benchmark_header_row}']
    benchmark_header_cell.value = "Benchmarks Return are all in Domestic Currencies."
    benchmark_header_cell.font = Font(name='Aptos', size=11, bold=True, color='FF0000')
    benchmark_header_cell.alignment = Alignment(horizontal='left', vertical='center')
    benchmark_header_cell.border = no_border
    
    benchmark_start_row = benchmark_header_row + 1
    
    # Define benchmark colors
    benchmark_colors = {
        "USD/CAD": "fcd5b4",
        "S&P 500": "b8cce4",
        "Dow Jones": "b8cce4",
        "Nasdaq": "b8cce4",
        "ACWI": "b8cce4",
        "TSX60": "d8e4bc"
    }
    
    for bench_name in BENCHMARK_ORDER:
        bench_row = benchmark_start_row
        bench_ticker_cell = ws[f'B{bench_row}']
        bench_ticker_cell.value = bench_name
        bench_ticker_cell.font = aptos_font_bold
        bench_ticker_cell.border = no_border
        bench_ticker_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        bench_color = benchmark_colors.get(bench_name, "FFFFFF")
        bench_fill = PatternFill(start_color=bench_color, end_color=bench_color, fill_type='solid')
        bench_ticker_cell.fill = bench_fill
        
        for period_idx in range(12):  # All 12 months
            col_idx = period_col_mapping[period_idx]
            # Get benchmark return if period exists, otherwise use 0.0
            has_data = period_idx < len(monthly_periods)
            if has_data:
                period = monthly_periods[period_idx]
                bench_return = monthly_benchmark_returns[bench_name].get(period, 0.0)
            else:
                bench_return = 0.0
            
            # Return column
            return_cell = ws[f'{get_column_letter(col_idx)}{bench_row}']
            if has_data:
                return_cell.value = bench_return
                return_cell.number_format = '0.00%;(0.00%)'
                return_cell.alignment = Alignment(horizontal='right', vertical='center')
                return_cell.fill = bench_fill
                if bench_return >= 0:
                    return_cell.font = Font(name='Aptos', size=11, color='006100')
                else:
                    return_cell.font = Font(name='Aptos', size=11, color='C00000')
            else:
                return_cell.value = "No Data"
                return_cell.number_format = '@'  # Text format
                return_cell.alignment = Alignment(horizontal='center', vertical='center')
                return_cell.font = Font(name='Aptos', size=11, color='808080')  # Grey font
                return_cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # Light grey background
            return_cell.border = no_border
            
            # Contrib column (empty)
            ws[f'{get_column_letter(col_idx + 1)}{bench_row}'].value = ""
            ws[f'{get_column_letter(col_idx + 1)}{bench_row}'].border = no_border
            ws[f'{get_column_letter(col_idx + 1)}{bench_row}'].fill = bench_fill
        
        # YTD Return
        first_date = pd.to_datetime(dates[0], format="%d/%m/%Y")
        last_date = pd.to_datetime(dates[-1], format="%d/%m/%Y")
        
        ticker = BENCHMARK_TICKERS[bench_name]
        if ticker == FX_TICKER:
            from market_data import get_price_on_date
            fx_start = get_price_on_date(FX_TICKER, first_date, cache)
            fx_end = get_price_on_date(FX_TICKER, last_date, cache)
            ytd_bench_return = (fx_end / fx_start) - 1
        else:
            from market_data import get_price_on_date
            price_start = get_price_on_date(ticker, first_date, cache)
            price_end = get_price_on_date(ticker, last_date, cache)
            ytd_bench_return = (price_end / price_start) - 1
        
        ytd_return_cell = ws[f'{get_column_letter(ytd_col)}{bench_row}']
        ytd_return_cell.value = ytd_bench_return
        ytd_return_cell.number_format = '0.00%;(0.00%)'
        ytd_return_cell.alignment = Alignment(horizontal='right', vertical='center')
        ytd_return_cell.border = no_border
        ytd_return_cell.font = aptos_font
        ytd_return_cell.fill = bench_fill
        if ytd_bench_return >= 0:
            ytd_return_cell.font = Font(name='Aptos', size=11, color='006100')
        else:
            ytd_return_cell.font = Font(name='Aptos', size=11, color='C00000')
        
        ws[f'{get_column_letter(ytd_col + 1)}{bench_row}'].value = ""
        ws[f'{get_column_letter(ytd_col + 1)}{bench_row}'].border = no_border
        ws[f'{get_column_letter(ytd_col + 1)}{bench_row}'].fill = bench_fill
        
        benchmark_start_row += 1
    
    # Set column widths
    max_col = max([ytd_col + 1] + [period_col_mapping[idx] + 3 for idx in period_col_mapping])
    
    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        
        if col == 1:  # Column A - empty border column (10px)
            ws.column_dimensions[col_letter].width = 10 / 7  # 10px border column
        elif col == 2:  # Ticker column (Column B)
            ws.column_dimensions[col_letter].width = 15
        elif col == 3:  # Spacer column (Column C)
            ws.column_dimensions[col_letter].width = 20 / 7  # 10px spacer column
        elif col >= col_offset and col < ytd_col:
            is_spacer = False
            for period_idx, mapped_col in period_col_mapping.items():
                if col == mapped_col + 2:  # Spacer is after Return/Contrib
                    is_spacer = True
                    break
            
            if is_spacer:
                ws.column_dimensions[col_letter].width = 20 / 7
            else:
                for period_idx, mapped_col in period_col_mapping.items():
                    if mapped_col <= col < mapped_col + 2:  # Return and Contrib columns
                        ws.column_dimensions[col_letter].width = 65 / 7
                        break
        elif col >= ytd_col:
            ws.column_dimensions[col_letter].width = 100 / 7
        else:
            ws.column_dimensions[col_letter].width = 12
    
    ws.sheet_view.showGridLines = False
    
    # Apply borders (similar to period sheet)
    header_row = data_start_row - 2  # Month header row
    subheader_row = data_start_row - 1  # Subheader row
    
    for period_idx in range(12):  # All 12 months
        col_idx = period_col_mapping[period_idx]
        end_col = col_idx + 1  # Only Return and Contrib (no Weight)
        
        # Thick border on date span merged cell
        for col in range(col_idx, end_col + 1):
            cell = ws[f'{get_column_letter(col)}{header_row}']
            cell.border = Border(
                left=Side(style='medium', color='000000') if col == col_idx else None,
                right=Side(style='medium', color='000000') if col == end_col else None,
                top=Side(style='medium', color='000000'),
                bottom=Side(style='medium', color='000000')
            )
        
        # Top border (subheader row)
        for col in range(col_idx, end_col + 1):
            cell = ws[f'{get_column_letter(col)}{subheader_row}']
            cell.border = Border(
                left=Side(style='medium', color='000000') if col == col_idx else None,
                right=Side(style='medium', color='000000') if col == end_col else None,
                top=Side(style='medium', color='000000'),
                bottom=None
            )
        
        # Left and right borders for data rows
        for row in range(data_start_row, holdings_end_row + 1):
            left_cell = ws[f'{get_column_letter(col_idx)}{row}']
            current_border = left_cell.border
            new_left_border = Border(
                left=Side(style='medium', color='000000'),
                right=current_border.right if current_border and current_border.right else None,
                top=Side(style='medium', color='000000') if row == data_start_row else None,
                bottom=Side(style='medium', color='000000') if row == holdings_end_row else None
            )
            left_cell.border = new_left_border
            
            right_cell = ws[f'{get_column_letter(end_col)}{row}']
            current_border = right_cell.border
            new_right_border = Border(
                left=current_border.left if current_border and current_border.left else None,
                right=Side(style='medium', color='000000'),
                top=Side(style='medium', color='000000') if row == data_start_row else None,
                bottom=Side(style='medium', color='000000') if row == holdings_end_row else None
            )
            right_cell.border = new_right_border
        
        # Bottom border
        for col in range(col_idx, end_col + 1):
            cell = ws[f'{get_column_letter(col)}{holdings_end_row}']
            current_border = cell.border
            cell.border = Border(
                left=Side(style='medium', color='000000') if col == col_idx else (current_border.left if current_border and current_border.left else None),
                right=Side(style='medium', color='000000') if col == end_col else (current_border.right if current_border and current_border.right else None),
                top=current_border.top if current_border and current_border.top else None,
                bottom=Side(style='medium', color='000000')
            )
    
    # YTD borders
    ytd_end_col = ytd_col + 1
    for col in range(ytd_col, ytd_end_col + 1):
        cell = ws[f'{get_column_letter(col)}{header_row}']
        cell.border = Border(
            left=Side(style='medium', color='000000') if col == ytd_col else None,
            right=Side(style='medium', color='000000') if col == ytd_end_col else None,
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
    
    for col in range(ytd_col, ytd_end_col + 1):
        cell = ws[f'{get_column_letter(col)}{subheader_row}']
        cell.border = Border(
            left=Side(style='medium', color='000000') if col == ytd_col else None,
            right=Side(style='medium', color='000000') if col == ytd_end_col else None,
            top=Side(style='medium', color='000000'),
            bottom=None
        )
    
    for row in range(data_start_row, holdings_end_row + 1):
        left_cell = ws[f'{get_column_letter(ytd_col)}{row}']
        current_border = left_cell.border
        new_left_border = Border(
            left=Side(style='medium', color='000000'),
            right=current_border.right if current_border and current_border.right else None,
            top=Side(style='medium', color='000000') if row == data_start_row else None,
            bottom=Side(style='medium', color='000000') if row == holdings_end_row else None
        )
        left_cell.border = new_left_border
        
        right_cell = ws[f'{get_column_letter(ytd_end_col)}{row}']
        current_border = right_cell.border
        new_right_border = Border(
            left=current_border.left if current_border and current_border.left else None,
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000') if row == data_start_row else None,
            bottom=Side(style='medium', color='000000') if row == holdings_end_row else None
        )
        right_cell.border = new_right_border
    
    for col in range(ytd_col, ytd_end_col + 1):
        cell = ws[f'{get_column_letter(col)}{holdings_end_row}']
        current_border = cell.border
        cell.border = Border(
            left=Side(style='medium', color='000000') if col == ytd_col else (current_border.left if current_border and current_border.left else None),
            right=Side(style='medium', color='000000') if col == ytd_end_col else (current_border.right if current_border and current_border.right else None),
            top=current_border.top if current_border and current_border.top else None,
            bottom=Side(style='medium', color='000000')
        )
    
    # Apply thick borders around ticker column
    for row in range(data_start_row, holdings_end_row + 1):
        ticker_cell = ws[f'B{row}']
        top_border = Side(style='medium', color='000000') if row == data_start_row else None
        bottom_border = Side(style='medium', color='000000') if row == holdings_end_row else None
        ticker_cell.border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=top_border,
            bottom=bottom_border
        )
    
    # Set print settings - center vertically and horizontally
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    
    # Set page setup to fit all rows in one page
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = False
    try:
        delattr(ws.page_setup, 'scale')
    except AttributeError:
        pass
    
    # Set custom header and footer
    ws.oddHeader.center.text = '&"Aptos,Bold"&26&[Tab]'
    ws.oddFooter.center.text = '&"Aptos,Bold"&12Page &[Page] of &[Pages]'

