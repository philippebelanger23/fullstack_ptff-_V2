"""Top Contributors/Disruptors sheet creation."""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
import pandas as pd
from constants import CASH_TICKER, BENCHMARK_TICKERS


def group_periods_by_month(periods):
    """Group periods by calendar month."""
    month_groups = {}
    for period_idx, period in enumerate(periods):
        period_start, period_end = period
        # Use period end to determine which month it belongs to
        month_key = (period_end.year, period_end.month)
        if month_key not in month_groups:
            month_groups[month_key] = []
        month_groups[month_key].append((period_idx, period))
    return month_groups


def group_months_by_quarter(sorted_months):
    """Group months into quarters (3 months + 1 quarter)."""
    quarters = []
    for i in range(0, len(sorted_months), 3):
        quarter_months = sorted_months[i:i+3]
        if len(quarter_months) == 3:
            quarters.append((quarter_months, True))  # True = has quarter column
        else:
            quarters.append((quarter_months, False))  # False = no quarter column
    return quarters


def aggregate_monthly_data(df, periods, month_periods):
    """Aggregate data for a month by summing contributions from all periods in that month."""
    month_data = []
    
    for _, row_data in df.iterrows():
        ticker = row_data['Ticker']
        
        # Skip benchmarks only (include cash)
        if ticker in BENCHMARK_TICKERS.values():
            continue
        
        total_weight = 0.0
        total_contrib = 0.0
        weighted_return_sum = 0.0
        weight_sum_for_return = 0.0
        most_recent_weight = 0.0  # Track most recent weight (closest to end of period)
        
        # Sort periods by end date to get the most recent one
        # month_periods is list of (period_idx, period) where period is (start_date, end_date)
        sorted_periods = sorted(month_periods, key=lambda x: x[1][1])  # Sort by period end date
        
        # Handle cash separately (return and contrib are always 0)
        if ticker == CASH_TICKER:
            # Get most recent weight for cash
            if sorted_periods:
                last_period_idx = sorted_periods[-1][0]
                most_recent_weight = row_data.get(f'Weight_{last_period_idx}', 0.0)
            else:
                # Fallback: use max weight if no sorted periods
                for period_idx, period in month_periods:
                    weight = row_data.get(f'Weight_{period_idx}', 0.0)
                    most_recent_weight = max(most_recent_weight, weight)
            
            # Cash always has 0 return and 0 contribution
            month_data.append({
                'Ticker': ticker,
                'Weight': most_recent_weight,
                'Return': 0.0,
                'Contrib': 0.0
            })
            continue
        
        # Aggregate data from all periods in this month for non-cash tickers
        for period_idx, period in month_periods:
            weight = row_data.get(f'Weight_{period_idx}', 0.0)
            return_val = row_data.get(f'Return_{period_idx}', 0.0)
            contrib = row_data.get(f'Contrib_{period_idx}', 0.0)
            
            # Sum contributions
            total_contrib += contrib
            
            # Calculate weighted return (sum of weight * return)
            if weight > 0:
                weighted_return_sum += return_val * weight
                weight_sum_for_return += weight
        
        # Get most recent weight (from last period in sorted list)
        if sorted_periods:
            last_period_idx = sorted_periods[-1][0]
            most_recent_weight = row_data.get(f'Weight_{last_period_idx}', 0.0)
        else:
            # Fallback: use max weight if no sorted periods
            for period_idx, period in month_periods:
                weight = row_data.get(f'Weight_{period_idx}', 0.0)
                most_recent_weight = max(most_recent_weight, weight)
        
        total_weight = most_recent_weight
        
        # Calculate average weighted return for the month
        avg_return = weighted_return_sum / weight_sum_for_return if weight_sum_for_return > 0 else 0.0
        
        # Include all tickers (even if weight is 0) to ensure weights sum correctly
        month_data.append({
            'Ticker': ticker,
            'Weight': total_weight,
            'Return': avg_return,
            'Contrib': total_contrib
        })
    
    return pd.DataFrame(month_data)


def aggregate_quarterly_data(df, periods, quarter_months, month_groups):
    """Aggregate data for a quarter by combining all months in the quarter."""
    quarter_data = []
    
    # Get all period indices for all months in the quarter
    all_quarter_periods = []
    for (year, month) in quarter_months:
        if (year, month) in month_groups:
            all_quarter_periods.extend(month_groups[(year, month)])
    
    if not all_quarter_periods:
        return pd.DataFrame()
    
    for _, row_data in df.iterrows():
        ticker = row_data['Ticker']
        
        # Skip benchmarks only (include cash)
        if ticker in BENCHMARK_TICKERS.values():
            continue
        
        total_weight = 0.0
        total_contrib = 0.0
        weighted_return_sum = 0.0
        weight_sum_for_return = 0.0
        
        # Handle cash separately (return and contrib are always 0)
        if ticker == CASH_TICKER:
            # Get most recent weight for cash (from last period in quarter)
            most_recent_weight = 0.0
            sorted_quarter_periods = sorted(all_quarter_periods, key=lambda x: x[1][1])
            if sorted_quarter_periods:
                last_period_idx = sorted_quarter_periods[-1][0]
                most_recent_weight = row_data.get(f'Weight_{last_period_idx}', 0.0)
            else:
                # Fallback: use max weight
                for period_idx, period in all_quarter_periods:
                    weight = row_data.get(f'Weight_{period_idx}', 0.0)
                    most_recent_weight = max(most_recent_weight, weight)
            
            # Cash always has 0 return and 0 contribution
            quarter_data.append({
                'Ticker': ticker,
                'Weight': most_recent_weight,
                'Return': 0.0,
                'Contrib': 0.0
            })
            continue
        
        # Aggregate data from all periods in the quarter for non-cash tickers
        # Sort periods by end date to get the most recent one
        sorted_quarter_periods = sorted(all_quarter_periods, key=lambda x: x[1][1])  # Sort by period end date
        
        for period_idx, period in all_quarter_periods:
            weight = row_data.get(f'Weight_{period_idx}', 0.0)
            return_val = row_data.get(f'Return_{period_idx}', 0.0)
            contrib = row_data.get(f'Contrib_{period_idx}', 0.0)
            
            # Sum contributions
            total_contrib += contrib
            
            # Calculate weighted return (sum of weight * return)
            if weight > 0:
                weighted_return_sum += return_val * weight
                weight_sum_for_return += weight
        
        # Get most recent weight (from last period in sorted list)
        if sorted_quarter_periods:
            last_period_idx = sorted_quarter_periods[-1][0]
            most_recent_weight = row_data.get(f'Weight_{last_period_idx}', 0.0)
        else:
            # Fallback: use max weight if no sorted periods
            for period_idx, period in all_quarter_periods:
                weight = row_data.get(f'Weight_{period_idx}', 0.0)
                most_recent_weight = max(most_recent_weight, weight)
        
        total_weight = most_recent_weight
        
        # Calculate average weighted return for the quarter
        avg_return = weighted_return_sum / weight_sum_for_return if weight_sum_for_return > 0 else 0.0
        
        # Include all tickers (even if weight is 0) to ensure weights sum correctly
        quarter_data.append({
            'Ticker': ticker,
            'Weight': total_weight,
            'Return': avg_return,
            'Contrib': total_contrib
        })
    
    return pd.DataFrame(quarter_data)


def create_table(ws, table_start_col, start_row, month_name, month_df, is_quarter=False):
    """Create a single table for a month or quarter."""
    table_width = 4  # Ticker, Weight, Performance, Contributions
    current_row = start_row
    
    # Define fonts
    aptos_font = Font(name='Aptos', size=11)
    aptos_font_bold = Font(name='Aptos', size=11, bold=True)
    
    # Medium border for table outline (matching period_sheet and monthly_sheet)
    medium_side = Side(style='medium', color='000000')
    
    # Find top 5 contributors and top 5 disruptors
    top_contributors = month_df.nlargest(5, 'Contrib')
    top_disruptors = month_df.nsmallest(5, 'Contrib')
    
    # Month/Quarter header - 2 rows wide, merged and centered
    month_header_col = table_start_col
    month_header_end_col = table_start_col + table_width - 1
    month_header_start_row = current_row
    month_header_end_row = current_row + 1
    ws.merge_cells(f'{get_column_letter(month_header_col)}{current_row}:{get_column_letter(month_header_end_col)}{current_row + 1}')
    month_header_cell = ws[f'{get_column_letter(month_header_col)}{current_row}']
    month_header_cell.value = month_name
    month_header_cell.font = Font(name='Aptos', size=15, bold=True, color='FFFFFF')
    month_header_cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    month_header_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add complete border around title (merged cell spanning 2 rows)
    for col in range(month_header_col, month_header_end_col + 1):
        for row in range(month_header_start_row, month_header_end_row + 1):
            cell = ws[f'{get_column_letter(col)}{row}']
            cell.border = Border(
                left=medium_side if col == month_header_col else None,
                right=medium_side if col == month_header_end_col else None,
                top=medium_side if row == month_header_start_row else None,
                bottom=medium_side if row == month_header_end_row else None
            )
    
    current_row += 2
    
    # Empty row between title and Top Contributors
    empty_row = current_row
    current_row += 1
    
    # Add left and right borders to empty row
    for col in range(table_start_col, table_start_col + table_width):
        cell = ws[f'{get_column_letter(col)}{empty_row}']
        cell.border = Border(
            left=medium_side if col == table_start_col else None,
            right=medium_side if col == table_start_col + table_width - 1 else None,
            top=None,
            bottom=None
        )
    
    # Top Contributors header
    contrib_header_row = current_row
    contrib_header_col = table_start_col
    contrib_header_end_col = table_start_col + table_width - 1
    ws.merge_cells(f'{get_column_letter(contrib_header_col)}{current_row}:{get_column_letter(contrib_header_end_col)}{current_row}')
    contrib_header_cell = ws[f'{get_column_letter(contrib_header_col)}{current_row}']
    contrib_header_cell.value = "Top Contributors"
    contrib_header_cell.font = Font(name='Aptos', size=11, bold=True, color='000000')
    contrib_header_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add left and right borders to Top Contributors row
    for col in range(contrib_header_col, contrib_header_end_col + 1):
        cell = ws[f'{get_column_letter(col)}{contrib_header_row}']
        cell.border = Border(
            left=medium_side if col == contrib_header_col else None,
            right=medium_side if col == contrib_header_end_col else None,
            top=None,
            bottom=None
        )
    
    current_row += 1
    
    # Column headers for Top Contributors
    contrib_headers = ["Ticker", "Weight", "Performance", "Contrib. (bps)"]
    for col_idx, header in enumerate(contrib_headers):
        col = table_start_col + col_idx
        header_cell = ws[f'{get_column_letter(col)}{current_row}']
        header_cell.value = header
        header_cell.font = Font(name='Aptos', size=11, bold=True, color='FFFFFF')
        header_cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        # Outer borders on header row
        header_cell.border = Border(
            left=medium_side if col_idx == 0 else None,
            right=medium_side if col_idx == table_width - 1 else None,
            top=medium_side,
            bottom=None
        )
    
    current_row += 1
    
    # Data rows - Top Contributors
    contrib_data_start_row = current_row
    contrib_last_row = contrib_data_start_row + len(top_contributors) - 1
    for idx, (_, contrib_row) in enumerate(top_contributors.iterrows()):
        row = contrib_data_start_row + idx
        
        # Ticker
        ticker_cell = ws[f'{get_column_letter(table_start_col)}{row}']
        ticker_cell.value = contrib_row['Ticker']
        ticker_cell.font = aptos_font_bold
        ticker_cell.alignment = Alignment(horizontal='left', vertical='center')
        # Left border on all data rows
        ticker_cell.border = Border(
            left=medium_side,
            right=None,
            top=None,
            bottom=None
        )
        
        # Weight
        weight_cell = ws[f'{get_column_letter(table_start_col + 1)}{row}']
        weight_cell.value = contrib_row['Weight']
        weight_cell.number_format = '0.00%;(0.00%)'
        weight_cell.alignment = Alignment(horizontal='center', vertical='center')
        weight_cell.font = aptos_font
        weight_cell.border = Border()  # No borders
        
        # Performance
        perf_cell = ws[f'{get_column_letter(table_start_col + 2)}{row}']
        perf_cell.value = contrib_row['Return']
        perf_cell.number_format = '0.00%;(0.00%)'
        perf_cell.alignment = Alignment(horizontal='center', vertical='center')
        perf_cell.font = aptos_font
        if contrib_row['Return'] >= 0:
            perf_cell.font = Font(name='Aptos', size=11, color='006100')
        else:
            perf_cell.font = Font(name='Aptos', size=11, color='C00000')
        perf_cell.border = Border()  # No borders
        
        # Contributions (in basis points)
        contrib_cell = ws[f'{get_column_letter(table_start_col + 3)}{row}']
        contrib_bps = contrib_row['Contrib'] * 10000  # Convert to basis points
        contrib_cell.value = contrib_bps
        contrib_cell.number_format = '0;(0)'  # Format negative as (15) not -15
        contrib_cell.alignment = Alignment(horizontal='center', vertical='center')
        contrib_cell.font = Font(name='Aptos', size=11, bold=True, color='006100')
        # Right border on all data rows
        contrib_cell.border = Border(
            right=medium_side,
            left=None,
            top=None,
            bottom=None
        )
    
    # Add underline under last contributor ticker (thin border)
    if len(top_contributors) > 0:
        last_contrib_row = contrib_last_row
        for col_idx in range(table_start_col, table_start_col + table_width):
            cell = ws[f'{get_column_letter(col_idx)}{last_contrib_row}']
            current_border = cell.border
            cell.border = Border(
                left=current_border.left if current_border and current_border.left else None,
                right=current_border.right if current_border and current_border.right else None,
                top=current_border.top if current_border and current_border.top else None,
                bottom=Side(style='thin', color='000000')
            )
    
    # Summary row for Contributors
    contrib_summary_row = contrib_data_start_row + len(top_contributors)
    contrib_total_weight = top_contributors['Weight'].sum()
    contrib_total_contrib = top_contributors['Contrib'].sum()
    
    # Summary header (Σ)
    summary_ticker_cell = ws[f'{get_column_letter(table_start_col)}{contrib_summary_row}']
    summary_ticker_cell.value = "Σ"
    summary_ticker_cell.font = aptos_font_bold
    summary_ticker_cell.alignment = Alignment(horizontal='center', vertical='center')
    summary_ticker_cell.border = Border(
        left=medium_side,
        right=None,
        top=None,
        bottom=None  # Bottom border will be added by final border code
    )
    
    summary_weight_cell = ws[f'{get_column_letter(table_start_col + 1)}{contrib_summary_row}']
    summary_weight_cell.value = contrib_total_weight
    summary_weight_cell.number_format = '0.00%;(0.00%)'
    summary_weight_cell.alignment = Alignment(horizontal='center', vertical='center')
    summary_weight_cell.font = aptos_font_bold
    summary_weight_cell.border = Border()  # No borders
    
    summary_perf_cell = ws[f'{get_column_letter(table_start_col + 2)}{contrib_summary_row}']
    summary_perf_cell.value = ""
    summary_perf_cell.font = aptos_font_bold
    summary_perf_cell.border = Border()  # No borders
    
    summary_contrib_cell = ws[f'{get_column_letter(table_start_col + 3)}{contrib_summary_row}']
    contrib_total_bps = contrib_total_contrib * 10000
    summary_contrib_cell.value = contrib_total_bps
    summary_contrib_cell.number_format = '0;(0)'  # Format negative as (15) not -15
    summary_contrib_cell.alignment = Alignment(horizontal='center', vertical='center')
    summary_contrib_cell.font = Font(name='Aptos', size=11, bold=True, color='006100')
    summary_contrib_cell.border = Border(
        right=medium_side,
        left=None,
        top=None,
        bottom=None  # Bottom border will be added by final border code
    )
    
    # Empty row between sum row of top contributors and Top Disruptors title
    current_row = contrib_summary_row + 2
    
    # Top Disruptors header
    disrupt_header_col = table_start_col
    disrupt_header_end_col = table_start_col + table_width - 1
    ws.merge_cells(f'{get_column_letter(disrupt_header_col)}{current_row}:{get_column_letter(disrupt_header_end_col)}{current_row}')
    disrupt_header_cell = ws[f'{get_column_letter(disrupt_header_col)}{current_row}']
    disrupt_header_cell.value = "Top Disruptors"
    disrupt_header_cell.font = Font(name='Aptos', size=11, bold=True, color='000000')
    disrupt_header_cell.alignment = Alignment(horizontal='center', vertical='center')
    # No border on header - only on data cells
    current_row += 1
    
    # Column headers for Top Disruptors
    disrupt_headers = ["Ticker", "Weight", "Performance", "Contrib. (bps)"]
    for col_idx, header in enumerate(disrupt_headers):
        col = table_start_col + col_idx
        header_cell = ws[f'{get_column_letter(col)}{current_row}']
        header_cell.value = header
        header_cell.font = Font(name='Aptos', size=11, bold=True, color='FFFFFF')
        header_cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        # Outer borders on header row
        header_cell.border = Border(
            left=medium_side if col_idx == 0 else None,
            right=medium_side if col_idx == table_width - 1 else None,
            top=medium_side,
            bottom=None
        )
    
    current_row += 1
    
    # Data rows - Top Disruptors
    disrupt_data_start_row = current_row
    disrupt_last_row = disrupt_data_start_row + len(top_disruptors) - 1
    for idx, (_, disrupt_row) in enumerate(top_disruptors.iterrows()):
        row = disrupt_data_start_row + idx
        
        # Ticker
        ticker_cell = ws[f'{get_column_letter(table_start_col)}{row}']
        ticker_cell.value = disrupt_row['Ticker']
        ticker_cell.font = aptos_font_bold
        ticker_cell.alignment = Alignment(horizontal='left', vertical='center')
        # Left border on all data rows
        ticker_cell.border = Border(
            left=medium_side,
            right=None,
            top=None,
            bottom=None
        )
        
        # Weight
        weight_cell = ws[f'{get_column_letter(table_start_col + 1)}{row}']
        weight_cell.value = disrupt_row['Weight']
        weight_cell.number_format = '0.00%;(0.00%)'
        weight_cell.alignment = Alignment(horizontal='center', vertical='center')
        weight_cell.font = aptos_font
        weight_cell.border = Border()  # No borders
        
        # Performance
        perf_cell = ws[f'{get_column_letter(table_start_col + 2)}{row}']
        perf_cell.value = disrupt_row['Return']
        perf_cell.number_format = '0.00%;(0.00%)'
        perf_cell.alignment = Alignment(horizontal='center', vertical='center')
        perf_cell.font = aptos_font
        if disrupt_row['Return'] >= 0:
            perf_cell.font = Font(name='Aptos', size=11, color='006100')
        else:
            perf_cell.font = Font(name='Aptos', size=11, color='C00000')
        perf_cell.border = Border()  # No borders
        
        # Contributions (in basis points)
        contrib_cell = ws[f'{get_column_letter(table_start_col + 3)}{row}']
        contrib_bps = disrupt_row['Contrib'] * 10000  # Convert to basis points
        contrib_cell.value = contrib_bps
        contrib_cell.number_format = '0;(0)'  # Format negative as (15) not -15
        contrib_cell.alignment = Alignment(horizontal='center', vertical='center')
        contrib_cell.font = Font(name='Aptos', size=11, bold=True, color='C00000')
        # Right border on all data rows
        contrib_cell.border = Border(
            right=medium_side,
            left=None,
            top=None,
            bottom=None
        )
    
    # Add underline under last disruptor ticker (thin border)
    if len(top_disruptors) > 0:
        last_disrupt_row = disrupt_last_row
        for col_idx in range(table_start_col, table_start_col + table_width):
            cell = ws[f'{get_column_letter(col_idx)}{last_disrupt_row}']
            current_border = cell.border
            cell.border = Border(
                left=current_border.left if current_border and current_border.left else None,
                right=current_border.right if current_border and current_border.right else None,
                top=current_border.top if current_border and current_border.top else None,
                bottom=Side(style='thin', color='000000')
            )
    
    # Summary row for Disruptors
    disrupt_summary_row = disrupt_data_start_row + len(top_disruptors)
    disrupt_total_weight = top_disruptors['Weight'].sum()
    disrupt_total_contrib = top_disruptors['Contrib'].sum()
    
    # Summary header (Σ) for Disruptors
    summary_ticker_cell = ws[f'{get_column_letter(table_start_col)}{disrupt_summary_row}']
    summary_ticker_cell.value = "Σ"
    summary_ticker_cell.font = aptos_font_bold
    summary_ticker_cell.alignment = Alignment(horizontal='center', vertical='center')
    summary_ticker_cell.border = Border(
        left=medium_side,
        right=None,
        top=None,
        bottom=None  # Bottom border will be added by final border code
    )
    
    summary_weight_cell = ws[f'{get_column_letter(table_start_col + 1)}{disrupt_summary_row}']
    summary_weight_cell.value = disrupt_total_weight
    summary_weight_cell.number_format = '0.00%;(0.00%)'
    summary_weight_cell.alignment = Alignment(horizontal='center', vertical='center')
    summary_weight_cell.font = aptos_font_bold
    summary_weight_cell.border = Border()  # No borders
    
    summary_perf_cell = ws[f'{get_column_letter(table_start_col + 2)}{disrupt_summary_row}']
    summary_perf_cell.value = ""
    summary_perf_cell.font = aptos_font_bold
    summary_perf_cell.border = Border()  # No borders
    
    summary_contrib_cell = ws[f'{get_column_letter(table_start_col + 3)}{disrupt_summary_row}']
    disrupt_total_bps = disrupt_total_contrib * 10000
    summary_contrib_cell.value = disrupt_total_bps
    summary_contrib_cell.number_format = '0;(0)'  # Format negative as (15) not -15
    summary_contrib_cell.alignment = Alignment(horizontal='center', vertical='center')
    summary_contrib_cell.font = Font(name='Aptos', size=11, bold=True, color='C00000')
    summary_contrib_cell.border = Border(
        right=medium_side,
        left=None,
        top=None,
        bottom=None  # Bottom border will be added by final border code
    )
    
    # Empty row between sum row of top disruptors and Other Holdings row
    current_row = disrupt_summary_row + 2
    
    # Other holdings row
    other_row = current_row
    # Get all tickers not in top 5 contributors or top 5 disruptors
    excluded_tickers = set(top_contributors['Ticker'].tolist() + top_disruptors['Ticker'].tolist())
    other_holdings = month_df[~month_df['Ticker'].isin(excluded_tickers)]
    
    other_weight = 0.0
    other_contrib = 0.0
    other_return = 0.0
    
    if not other_holdings.empty:
        # Sum weights of all tickers not in top 5 (using most recent weight from month_df)
        other_weight = other_holdings['Weight'].sum()
        # Sum contributions of all tickers not in top 5
        other_contrib = other_holdings['Contrib'].sum()
        # Calculate weighted return
        other_return = (other_holdings['Return'] * other_holdings['Weight']).sum() / other_weight if other_weight > 0 else 0.0
    
    # Always show Other Holding row
    other_ticker_cell = ws[f'{get_column_letter(table_start_col)}{other_row}']
    other_ticker_cell.value = "Other Holdings"
    other_ticker_cell.font = aptos_font
    other_ticker_cell.alignment = Alignment(horizontal='left', vertical='center')
    other_ticker_cell.border = Border(
        left=medium_side,
        right=None,
        top=None,
        bottom=None
    )
    
    other_weight_cell = ws[f'{get_column_letter(table_start_col + 1)}{other_row}']
    other_weight_cell.value = other_weight
    other_weight_cell.number_format = '0.00%;(0.00%)'
    other_weight_cell.alignment = Alignment(horizontal='center', vertical='center')
    other_weight_cell.font = aptos_font
    other_weight_cell.border = Border()  # No borders
    
    other_perf_cell = ws[f'{get_column_letter(table_start_col + 2)}{other_row}']
    other_perf_cell.value = other_return
    other_perf_cell.number_format = '0.00%;(0.00%)'
    other_perf_cell.alignment = Alignment(horizontal='center', vertical='center')
    other_perf_cell.font = aptos_font
    if other_return >= 0:
        other_perf_cell.font = Font(name='Aptos', size=11, color='006100')
    else:
        other_perf_cell.font = Font(name='Aptos', size=11, color='C00000')
    other_perf_cell.border = Border()  # No borders
    
    other_contrib_cell = ws[f'{get_column_letter(table_start_col + 3)}{other_row}']
    other_contrib_bps = other_contrib * 10000
    other_contrib_cell.value = other_contrib_bps
    other_contrib_cell.number_format = '0;(0)'  # Format negative as (15) not -15
    other_contrib_cell.alignment = Alignment(horizontal='center', vertical='center')
    if other_contrib >= 0:
        other_contrib_cell.font = Font(name='Aptos', size=11, bold=True, color='006100')
    else:
        other_contrib_cell.font = Font(name='Aptos', size=11, bold=True, color='C00000')
    other_contrib_cell.border = Border(
        right=medium_side,
        left=None,
        top=None,
        bottom=None
    )
    
    # Empty row between Other Holding and Total Portfolio
    total_row = other_row + 2
    # Calculate Total Portfolio weight as sum of ALL weights in month_df (ensures 100%)
    # This is the sum of Top 5 Contributors + Top 5 Disruptors + Other Holdings
    total_weight = month_df['Weight'].sum()  # Sum all weights from month_df
    total_contrib = month_df['Contrib'].sum()  # Sum all contributions from month_df
    total_return = (month_df['Return'] * month_df['Weight']).sum() / total_weight if total_weight > 0 else 0.0
    
    # Total Portfolio - light grey fill
    light_grey_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    
    total_ticker_cell = ws[f'{get_column_letter(table_start_col)}{total_row}']
    total_ticker_cell.value = "Total Portfolio"
    total_ticker_cell.font = aptos_font_bold
    total_ticker_cell.alignment = Alignment(horizontal='left', vertical='center')
    total_ticker_cell.fill = light_grey_fill
    total_ticker_cell.border = Border(
        left=medium_side,
        right=None,
        top=None,
        bottom=medium_side
    )
    
    total_weight_cell = ws[f'{get_column_letter(table_start_col + 1)}{total_row}']
    total_weight_cell.value = total_weight
    total_weight_cell.number_format = '0.00%;(0.00%)'
    total_weight_cell.alignment = Alignment(horizontal='center', vertical='center')
    total_weight_cell.font = aptos_font_bold
    total_weight_cell.fill = light_grey_fill
    total_weight_cell.border = Border(
        left=None,
        right=None,
        top=None,
        bottom=medium_side
    )
    
    total_perf_cell = ws[f'{get_column_letter(table_start_col + 2)}{total_row}']
    total_perf_cell.value = ""
    total_perf_cell.font = aptos_font_bold
    total_perf_cell.fill = light_grey_fill
    total_perf_cell.border = Border(
        left=None,
        right=None,
        top=None,
        bottom=medium_side
    )
    
    total_contrib_cell = ws[f'{get_column_letter(table_start_col + 3)}{total_row}']
    total_contrib_bps = total_contrib * 10000
    total_contrib_cell.value = total_contrib_bps
    total_contrib_cell.number_format = '0;(0)'  # Format negative as (15) not -15
    total_contrib_cell.alignment = Alignment(horizontal='center', vertical='center')
    total_contrib_cell.font = aptos_font_bold
    if total_contrib >= 0:
        total_contrib_cell.font = Font(name='Aptos', size=11, bold=True, color='006100')
    else:
        total_contrib_cell.font = Font(name='Aptos', size=11, bold=True, color='C00000')
    total_contrib_cell.fill = light_grey_fill
    total_contrib_cell.border = Border(
        right=medium_side,
        left=None,
        top=None,
        bottom=medium_side
    )
    
    # Add complete border around the whole table
    # Table starts from header row (Top Contributors column headers) to Total Portfolio row
    table_start_row = contrib_data_start_row - 1  # Column headers row
    table_end_row = total_row
    
    # Ensure complete outer border around entire table
    # Top border on header row (all columns)
    for col_idx in range(table_start_col, table_start_col + table_width):
        cell = ws[f'{get_column_letter(col_idx)}{table_start_row}']
        current_border = cell.border
        cell.border = Border(
            left=current_border.left if current_border and current_border.left else None,
            right=current_border.right if current_border and current_border.right else None,
            top=medium_side,
            bottom=current_border.bottom if current_border and current_border.bottom else None
        )
    
    # Left border on all rows (Ticker column)
    for row in range(table_start_row, table_end_row + 1):
        cell = ws[f'{get_column_letter(table_start_col)}{row}']
        current_border = cell.border
        cell.border = Border(
            left=medium_side,
            right=current_border.right if current_border and current_border.right else None,
            top=current_border.top if current_border and current_border.top else None,
            bottom=current_border.bottom if current_border and current_border.bottom else None
        )
    
    # Right border on all rows (Contrib column)
    for row in range(table_start_row, table_end_row + 1):
        cell = ws[f'{get_column_letter(table_start_col + table_width - 1)}{row}']
        current_border = cell.border
        cell.border = Border(
            left=current_border.left if current_border and current_border.left else None,
            right=medium_side,
            top=current_border.top if current_border and current_border.top else None,
            bottom=current_border.bottom if current_border and current_border.bottom else None
        )
    
    # Bottom border on Total Portfolio row (all columns)
    for col_idx in range(table_start_col, table_start_col + table_width):
        cell = ws[f'{get_column_letter(col_idx)}{table_end_row}']
        current_border = cell.border
        cell.border = Border(
            left=current_border.left if current_border and current_border.left else None,
            right=current_border.right if current_border and current_border.right else None,
            top=current_border.top if current_border and current_border.top else None,
            bottom=medium_side
        )
    
    return total_row


def create_top_contributors_sheet(wb, df, periods, dates):
    """Create the Top Contributors/Disruptors sheet."""
    ws = wb.create_sheet("Top Contributors & Disruptors")
    
    # Do not freeze columns
    
    # Set column A to 10px (border effect)
    ws.column_dimensions['A'].width = 10 / 7  # Convert pixels to Excel units (approximately)
    
    # Group periods by month
    month_groups = group_periods_by_month(periods)
    
    # Find the first January month
    sorted_months = sorted(month_groups.keys())
    first_january_idx = None
    for i, (year, month) in enumerate(sorted_months):
        if month == 1:  # January
            first_january_idx = i
            break
    
    if first_january_idx is None:
        first_january_idx = 0
    
    sorted_months = sorted_months[first_january_idx:]
    
    # Group months into quarters (3 months + 1 quarter)
    quarters = group_months_by_quarter(sorted_months)
    
    # Calculate table column positions
    # Each table is 4 columns wide (Ticker, Weight, Performance, Contrib)
    # Spacing between tables
    table_spacing = 1
    table_width = 4
    
    # Column widths
    ticker_col_width = 100 / 7  # Convert to Excel units
    data_col_width = 120 / 7   # Convert to Excel units
    
    # Set column widths dynamically based on table positions
    # Column A: 10px border
    ws.column_dimensions['A'].width = 10 / 7
    
    # Set row 1 height to 10px (empty border row)
    ws.row_dimensions[1].height = 10  # Height in points (1 point ≈ 1.33 pixels)
    
    # For each table position (4 tables per row), set widths
    for table_idx in range(4):
        base_col = 2 + table_idx * (table_width + table_spacing)
        # Ticker column
        ws.column_dimensions[get_column_letter(base_col)].width = ticker_col_width
        # Weight, Performance, Contrib columns
        for i in range(1, 4):
            ws.column_dimensions[get_column_letter(base_col + i)].width = data_col_width
        # Spacer column (if not last table)
        if table_idx < 3:
            ws.column_dimensions[get_column_letter(base_col + table_width)].width = table_spacing
    
    # Start row is 2 (row B), leaving row 1 empty as spacer
    start_base_row = 2
    
    # Create tables for each quarter row
    max_table_height = 0  # Track the maximum height of tables in current quarter
    quarter_start_row = start_base_row
    q2_end_row = None  # Track where Q2 ends
    
    for quarter_idx, (quarter_months, has_quarter) in enumerate(quarters):
        # Calculate starting row for this quarter
        if quarter_idx > 0:
            # Add spacing between quarters (use max height from previous quarter + some spacing)
            quarter_start_row = quarter_start_row + max_table_height + 4  # 4 rows spacing between quarters
        
        max_table_height = 0  # Reset for this quarter
        
        # Create 3 monthly tables + 1 quarterly table
        for table_idx in range(4):
            if table_idx < len(quarter_months):
                # Monthly table
                (year, month) = quarter_months[table_idx]
                month_periods = month_groups[(year, month)]
                month_name = pd.to_datetime(f"{year}-{month:02d}-01").strftime("%B %Y")
                month_df = aggregate_monthly_data(df, periods, month_periods)
                
                if not month_df.empty:
                    # Calculate table start column (B=2, then add spacing)
                    table_start_col = 2 + table_idx * (table_width + table_spacing)
                    end_row = create_table(ws, table_start_col, quarter_start_row, month_name, month_df)
                    # Track the maximum height of tables in this quarter
                    table_height = end_row - quarter_start_row + 1
                    max_table_height = max(max_table_height, table_height)
            
            elif table_idx == 3 and has_quarter:
                # Quarterly table
                quarter_name = f"Q{quarter_idx + 1}"
                quarter_df = aggregate_quarterly_data(df, periods, quarter_months, month_groups)
                
                if not quarter_df.empty:
                    # Calculate table start column for quarter (4th position)
                    table_start_col = 2 + 3 * (table_width + table_spacing)
                    end_row = create_table(ws, table_start_col, quarter_start_row, quarter_name, quarter_df, is_quarter=True)
                    # Track the maximum height of tables in this quarter
                    table_height = end_row - quarter_start_row + 1
                    max_table_height = max(max_table_height, table_height)
        
        # After Q2 ends, store the end row for page break
        if quarter_idx == 1:  # Q2 just finished
            q2_end_row = quarter_start_row + max_table_height - 1
        
        # Insert page break right after Q2 ends (before Q3 starts)
        if quarter_idx == 2 and q2_end_row is not None:  # Q3 starts at quarter_idx 2
            # Add horizontal page break after the second empty row (out of 4 empty rows)
            # This ensures Q1/Q2 print on page 1 and Q3/Q4 start cleanly on page 2
            ws.row_breaks.append(Break(id=q2_end_row + 2))
    
    ws.sheet_view.showGridLines = False
    # Set print settings
    ws.page_setup.orientation = 'landscape'
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3
    # Set page setup to fit all columns in one page
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False
    try:
        delattr(ws.page_setup, 'scale')
    except AttributeError:
        pass
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    
    # Set custom header and footer
    ws.oddHeader.center.text = '&"Aptos,Bold"&48&[Tab]'
    ws.oddFooter.center.text = '&"Aptos,Bold"&18Page &[Page] of &[Pages]'